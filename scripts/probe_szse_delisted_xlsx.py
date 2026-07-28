#!/usr/bin/env python3
"""Inspect official SZSE 1793_ssgs/tab2 XLSX with board filters."""
from __future__ import annotations

import json
import re
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
BASE = "https://www.szse.cn/api/report/ShowReport"
REFERER = "https://www.szse.cn/"


def norm(v):
    return re.sub(r"\s+", "", "" if v is None else str(v)).strip()


def fetch(select_module: str | None):
    params = {"SHOWTYPE": "xlsx", "CATALOGID": "1793_ssgs", "TABKEY": "tab2"}
    if select_module is not None:
        params["selectModule"] = select_module
    url = BASE + "?" + urlencode(params)
    r = requests.get(url, headers={"User-Agent": UA, "Referer": REFERER}, timeout=45)
    r.raise_for_status()
    return url, r.content


def inspect(raw: bytes):
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    result = {"sheets": []}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        result["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "first_30": [[norm(x) for x in row] for row in rows[:30]],
        })
    return result


def main():
    out = {}
    for value in (None, "main", "nm"):
        key = "none" if value is None else value
        try:
            url, raw = fetch(value)
            out[key] = {"url": url, "bytes": len(raw), "xlsx": inspect(raw)}
            Path(f"data/lifecycle_api_probe/szse_delisted_{key}.xlsx").write_bytes(raw)
        except Exception as exc:
            out[key] = {"error": f"{type(exc).__name__}: {exc}"}
    Path("data/lifecycle_api_probe/szse_delisted_xlsx_report.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
