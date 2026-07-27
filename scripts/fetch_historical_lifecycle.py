#!/usr/bin/env python3
"""Build official historical main-board A-share lifecycle evidence.

Sources:
- SSE official stock-list SQL, STOCK_TYPE=1, COMPANY_STATUS=3 for delisted A shares.
- SSE same SQL without status filter as a state-partition control.
- SZSE official 1793_ssgs/tab2 XLSX, selectModule=main, for terminated main-board securities.
- CNINFO official code-change implementation notices for security-code identity transitions.

The SZSE main-board workbook contains both A and B shares. A shares are classified with
the exchange's official code-range rule; B shares are explicitly excluded from project scope.
"""
from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
SSE_SQL = "COMMON_SSE_CP_GPJCTPZ_GPLB_GP_L"
SSE_REFERER = "https://www.sse.com.cn/"
SZSE_REFERER = "https://www.szse.cn/"
CNINFO_REFERER = "https://www.cninfo.com.cn/"
TRANSITIONS = ROOT / "config/security_code_transitions.json"
SZSE_DELIST_URL = (
    "https://www.szse.cn/api/report/ShowReport?"
    + urlencode({"SHOWTYPE": "xlsx", "CATALOGID": "1793_ssgs", "TABKEY": "tab2", "selectModule": "main"})
)

EVENT_FIELDS = [
    "exchange", "code", "event_type", "effective_date", "announced_at", "name",
    "source_url", "source_sha256", "evidence_class", "source_payload",
]


@dataclass(frozen=True)
class DelistedSecurity:
    exchange: str
    code: str
    name: str
    list_date: str
    delist_date: str
    source_url: str
    source_sha256: str
    source_payload: str


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def clean_text(value: object) -> str:
    s = html.unescape("" if value is None else str(value))
    s = re.sub(r"<[^>]*>", "", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_date(value: object) -> str:
    s = clean_text(value)
    if not s:
        raise ValueError("empty date")
    if re.fullmatch(r"\d{8}", s):
        s = f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return date.fromisoformat(s).isoformat()


def parse_jsonp(raw: bytes) -> dict:
    s = raw.decode("utf-8", errors="strict").strip()
    m = re.match(r"^[^(]+\((.*)\)\s*;?$", s, flags=re.S)
    if m:
        s = m.group(1)
    obj = json.loads(s)
    if not isinstance(obj, dict):
        raise ValueError("SSE payload is not an object")
    return obj


def http_get(url: str, referer: str, attempts: int = 5) -> bytes:
    """Fetch primary evidence with bounded retries; invalid evidence still fails closed."""
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            r = requests.get(
                url,
                headers={
                    "User-Agent": UA,
                    "Referer": referer,
                    "Accept": "*/*",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
                    "Connection": "close",
                },
                timeout=60,
            )
            r.raise_for_status()
            if not r.content:
                raise RuntimeError(f"empty response: {url}")
            return r.content
        except (requests.RequestException, RuntimeError) as exc:
            last_error = exc
            if attempt >= attempts:
                break
            time.sleep(min(0.5 * (2 ** (attempt - 1)), 4.0))
    assert last_error is not None
    raise last_error


def sse_url(status: str) -> str:
    params = {
        "jsonCallBack": "cb_lifecycle_123", "STOCK_TYPE": "1", "REG_PROVINCE": "",
        "CSRC_CODE": "", "STOCK_CODE": "", "sqlId": SSE_SQL,
        "COMPANY_STATUS": status, "type": "inParams", "isPagination": "true",
        "pageHelp.cacheSize": "1", "pageHelp.beginPage": "1", "pageHelp.pageSize": "5000",
        "pageHelp.pageNo": "1", "pageHelp.endPage": "1",
    }
    return "https://query.sse.com.cn/sseQuery/commonQuery.do?" + urlencode(params)


def sse_rows(raw: bytes) -> list[dict]:
    payload = parse_jsonp(raw)
    rows = payload.get("result") or payload.get("data") or []
    if not isinstance(rows, list):
        raise ValueError("SSE result is not a list")
    return [r for r in rows if isinstance(r, dict)]


def fetch_sse_delisted(raw_dir: Path, current_codes: set[str]) -> tuple[list[DelistedSecurity], dict]:
    url_delisted = sse_url("3"); url_all = sse_url("")
    raw_delisted = http_get(url_delisted, SSE_REFERER); raw_all = http_get(url_all, SSE_REFERER)
    raw_dir.mkdir(parents=True, exist_ok=True)
    (raw_dir / "sse_main_a_status3_delisted.jsonp").write_bytes(raw_delisted)
    (raw_dir / "sse_main_a_all_statuses.jsonp").write_bytes(raw_all)
    digest = sha256(raw_delisted); out: list[DelistedSecurity] = []
    for row in sse_rows(raw_delisted):
        code = clean_text(row.get("A_STOCK_CODE"))
        if not re.fullmatch(r"\d{6}", code): raise ValueError(f"SSE delisted row has invalid code: {code!r}")
        if clean_text(row.get("LIST_BOARD")) not in ("", "1"): raise ValueError(f"SSE non-main row in STOCK_TYPE=1 delisted source: {code}")
        delist_raw = clean_text(row.get("DELIST_DATE"))
        if not delist_raw or delist_raw == "-": raise ValueError(f"SSE status=3 row lacks DELIST_DATE: {code}")
        out.append(DelistedSecurity("SSE",code,clean_text(row.get("COMPANY_ABBR") or row.get("SEC_NAME_CN")),normalize_date(row.get("LIST_DATE") or row.get("A_LIST_DATE")),normalize_date(delist_raw),url_delisted,digest,json.dumps(row,ensure_ascii=False,sort_keys=True)))
    delisted_codes={x.code for x in out}
    if len(delisted_codes)!=len(out): raise ValueError("duplicate SSE delisted codes")
    if delisted_codes & current_codes: raise ValueError(f"SSE current/delisted overlap: {sorted(delisted_codes & current_codes)[:20]}")
    all_codes={clean_text(r.get("A_STOCK_CODE")) for r in sse_rows(raw_all) if re.fullmatch(r"\d{6}",clean_text(r.get("A_STOCK_CODE")))}
    expected=current_codes|delisted_codes
    if all_codes!=expected: raise ValueError(f"SSE state partition mismatch: all={len(all_codes)} current={len(current_codes)} delisted={len(delisted_codes)} only_all={sorted(all_codes-expected)[:20]} only_partition={sorted(expected-all_codes)[:20]}")
    return sorted(out,key=lambda x:x.code),{"delisted_count":len(out),"current_count":len(current_codes),"all_status_count":len(all_codes),"state_partition_exact":True,"delisted_source_sha256":digest,"all_status_source_sha256":sha256(raw_all),"delisted_url":url_delisted,"all_status_url":url_all}


def szse_security_class(code: str) -> str:
    if not re.fullmatch(r"\d{6}",code): return "UNKNOWN"
    n=int(code)
    if 1<=n<=4999:
        if 1001<=n<=1199:return "MAIN_CDR"
        return "MAIN_A"
    if code.startswith("200"):return "MAIN_B"
    return "UNKNOWN"


def parse_szse_delisted_xlsx(raw: bytes, source_url: str) -> tuple[list[DelistedSecurity], dict]:
    wb=load_workbook(BytesIO(raw),read_only=False,data_only=True);ws=wb.active;rows=list(ws.iter_rows(values_only=True))
    if not rows:raise ValueError("SZSE delisted workbook empty")
    header=[clean_text(x) for x in rows[0]];expected=["证券代码","证券简称","上市日期","终止上市日期"]
    if header[:4]!=expected:raise ValueError(f"unexpected SZSE delisted header: {header}")
    digest=sha256(raw);out=[];class_counts={};unknown=[]
    for values in rows[1:]:
        if not any(v not in (None,"") for v in values):continue
        code=clean_text(values[0]);cls=szse_security_class(code);class_counts[cls]=class_counts.get(cls,0)+1
        if cls=="UNKNOWN":unknown.append(code);continue
        if cls!="MAIN_A":continue
        payload={"证券代码":code,"证券简称":clean_text(values[1]),"上市日期":clean_text(values[2]),"终止上市日期":clean_text(values[3]),"security_class":cls}
        out.append(DelistedSecurity("SZSE",code,payload["证券简称"],normalize_date(values[2]),normalize_date(values[3]),source_url,digest,json.dumps(payload,ensure_ascii=False,sort_keys=True)))
    if unknown:raise ValueError(f"unknown SZSE main-board security codes: {unknown[:20]}")
    codes={x.code for x in out}
    if len(codes)!=len(out):raise ValueError("duplicate SZSE delisted A-share codes")
    return sorted(out,key=lambda x:x.code),{"workbook_rows_ex_header":len(rows)-1,"class_counts":class_counts,"delisted_a_count":len(out),"source_sha256":digest,"source_url":source_url}


def fetch_szse_delisted(raw_dir: Path, current_codes: set[str]) -> tuple[list[DelistedSecurity], dict]:
    raw=http_get(SZSE_DELIST_URL,SZSE_REFERER);raw_dir.mkdir(parents=True,exist_ok=True);(raw_dir/"szse_main_delisted_selectModule_main.xlsx").write_bytes(raw)
    out,control=parse_szse_delisted_xlsx(raw,SZSE_DELIST_URL);delisted_codes={x.code for x in out}
    if delisted_codes & current_codes:raise ValueError(f"SZSE current/delisted overlap: {sorted(delisted_codes & current_codes)[:20]}")
    control["current_count"]=len(current_codes);control["current_delisted_disjoint"]=True;return out,control


def load_code_transitions() -> list[dict[str,str]]:
    raw=json.loads(TRANSITIONS.read_text(encoding="utf-8"))
    if not isinstance(raw,list):raise ValueError("security code transition config is not a list")
    out=[]
    for i,t in enumerate(raw):
        if not isinstance(t,dict):raise ValueError(f"bad transition #{i}")
        for k in ("exchange","old_code","new_code","effective_date","source_url","source_sha256","evidence_class"):
            if not t.get(k):raise ValueError(f"transition #{i} missing {k}")
        date.fromisoformat(str(t["effective_date"]));out.append({k:str(v) for k,v in t.items()})
    return out


def fetch_code_transition_evidence(raw_dir: Path) -> list[dict[str,str]]:
    controls=[];raw_dir.mkdir(parents=True,exist_ok=True)
    for t in load_code_transitions():
        raw=http_get(t["source_url"],CNINFO_REFERER);actual=sha256(raw)
        if actual!=t["source_sha256"]:raise ValueError(f"code-transition source hash mismatch {t['old_code']}->{t['new_code']}: {actual}")
        filename=f"code_transition_{t['old_code']}_to_{t['new_code']}.pdf";(raw_dir/filename).write_bytes(raw)
        controls.append({"exchange":t["exchange"],"old_code":t["old_code"],"new_code":t["new_code"],"effective_date":t["effective_date"],"source_url":t["source_url"],"source_sha256":actual,"raw_file":filename,"evidence_class":t["evidence_class"]})
    return controls


def read_current_seed(path: Path) -> list[dict[str,str]]:
    with path.open(encoding="utf-8",newline="") as f:rows=list(csv.DictReader(f))
    if not rows:raise RuntimeError("current lifecycle seed is empty")
    return rows


def to_events(items: list[DelistedSecurity]) -> list[dict[str,str]]:
    events=[]
    for x in items:
        for event_type,effective_date in (("LIST",x.list_date),("DELIST",x.delist_date)):
            events.append({"exchange":x.exchange,"code":x.code,"event_type":event_type,"effective_date":effective_date,"announced_at":"","name":x.name,"source_url":x.source_url,"source_sha256":x.source_sha256,"evidence_class":"RETROSPECTIVE_PRIMARY","source_payload":x.source_payload})
    return events


def write_events(path: Path, rows: list[dict[str,str]]) -> None:
    seen=set()
    for row in rows:
        key=(row["exchange"],row["code"],row["event_type"],row["effective_date"])
        if key in seen:raise ValueError(f"duplicate lifecycle event: {key}")
        seen.add(key)
    rows=sorted(rows,key=lambda r:(r["exchange"],r["code"],r["effective_date"],r["event_type"]));path.parent.mkdir(parents=True,exist_ok=True)
    with path.open("w",encoding="utf-8",newline="") as f:w=csv.DictWriter(f,fieldnames=EVENT_FIELDS);w.writeheader();w.writerows(rows)


def main() -> int:
    lifecycle_dir=ROOT/"data/security_lifecycle";raw_dir=lifecycle_dir/"raw";current_seed=read_current_seed(lifecycle_dir/"current_list_seed.csv")
    current_sse={r["code"] for r in current_seed if r["exchange"]=="SSE"};current_szse={r["code"] for r in current_seed if r["exchange"]=="SZSE"}
    sse_delisted,sse_control=fetch_sse_delisted(raw_dir,current_sse);szse_delisted,szse_control=fetch_szse_delisted(raw_dir,current_szse);transition_control=fetch_code_transition_evidence(raw_dir)
    historical=to_events(sse_delisted)+to_events(szse_delisted);write_events(lifecycle_dir/"events.csv",current_seed+historical)
    g1_manifest=json.loads((ROOT/"data/current_master/manifest.json").read_text(encoding="utf-8"));coverage_end=clean_text(g1_manifest.get("szse",{}).get("as_of")) or "2026-07-24"
    base_intervals=len(current_seed)+len(sse_delisted)+len(szse_delisted)
    manifest={"version":"V3.2.19-g2-historical-lifecycle-identity-aware","status":"HISTORICAL_BACKFILL_COMPLETE_CANDIDATE","source_fetched_at_utc":datetime.now(timezone.utc).isoformat(),"coverage_start":"2015-01-01","coverage_end":coverage_end,"historical_backfill_complete":True,"scope":"SSE_MAIN_A + SZSE_MAIN_A","counts":{"current_active":len(current_seed),"sse_current_active":len(current_sse),"szse_current_active":len(current_szse),"sse_delisted_a_all_history":len(sse_delisted),"szse_delisted_a_all_history":len(szse_delisted),"delisted_a_all_history":len(sse_delisted)+len(szse_delisted),"lifecycle_events":len(current_seed)+2*(len(sse_delisted)+len(szse_delisted)),"code_transitions":len(transition_control),"expected_intervals":base_intervals+len(transition_control)},"controls":{"sse":sse_control,"szse":szse_control,"code_transitions":transition_control},"evidence_policy":{"historical_exchange_lists":"RETROSPECTIVE_PRIMARY","current_active_seed":"RETROSPECTIVE_PRIMARY","security_code_transitions":"POINT_IN_TIME_PRIMARY","point_in_time_claim":"Security identity follows the exchange code effective on each date; predecessor history is never retroactively relabelled to a successor code."}}
    (lifecycle_dir/"manifest.json").write_text(json.dumps(manifest,ensure_ascii=False,indent=2),encoding="utf-8");print(json.dumps(manifest,ensure_ascii=False,indent=2));return 0


if __name__=="__main__":sys.exit(main())
