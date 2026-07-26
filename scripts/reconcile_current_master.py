#!/usr/bin/env python3
"""Independent same-exchange reconciliation for Stage2 G1.

Primary row-level master:
- SSE: current SQL JSONP stock-list endpoint
- SZSE: paginated CATALOGID=1110 JSON

Independent exchange-owned controls:
- SSE: official real-time equity quote list; CPXXSubType=ASH means RMB-traded main-board stock
- SZSE: official A-share XLSX download with explicit board column

G1 only reconciles when primary and independent-control MAIN-A code sets match exactly.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from io import BytesIO
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import load_workbook

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
SSE_PAGE = "https://www.sse.com.cn/market/price/trends/"
SZSE_PAGE = "https://www.szse.cn/certificate/maind/"
SSE_CONTROL = (
    "https://yunhq.sse.com.cn:32042/v1/sh1/list/exchange/equity"
    "?select=code%2Cname%2Ccpxxsubtype%2Ccpxxprodusta&begin=0&end=5000"
)
SZSE_CONTROL = "https://www.szse.cn/api/report/ShowReport?SHOWTYPE=xlsx&CATALOGID=1110&TABKEY=tab1"


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def fetch(url: str, referer: str) -> bytes:
    s = requests.Session()
    s.headers.update({"User-Agent": UA, "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7"})
    try:
        s.get(referer, timeout=20)
    except requests.RequestException:
        pass
    r = s.get(
        url,
        headers={"Referer": referer, "Accept": "*/*", "X-Requested-With": "XMLHttpRequest"},
        timeout=45,
    )
    r.raise_for_status()
    if not r.content:
        raise RuntimeError(f"empty control response: {url}")
    return r.content


def read_primary_codes(path: Path, exchange: str) -> set[str]:
    with path.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    codes = {r["code"] for r in rows if r.get("exchange") == exchange and r.get("board") == "MAIN"}
    if not codes:
        raise RuntimeError(f"empty primary code set: {exchange}")
    return codes


def parse_sse_quote(raw: bytes) -> tuple[set[str], dict[str, object]]:
    payload = json.loads(raw.decode("utf-8"))
    rows = payload.get("list")
    total = int(payload.get("total") or 0)
    if not isinstance(rows, list):
        raise ValueError("SSE quote control list is not an array")
    if total <= 0:
        raise ValueError(f"invalid SSE quote total={total}")
    if len(rows) != total:
        raise RuntimeError(f"SSE quote control truncated: total={total}, rows={len(rows)}")

    subtype_counts: dict[str, int] = {}
    codes: set[str] = set()
    for row in rows:
        if not isinstance(row, list) or len(row) < 3:
            raise ValueError(f"malformed SSE quote row: {row!r}")
        code = str(row[0]).strip()
        subtype = str(row[2]).strip()
        subtype_counts[subtype] = subtype_counts.get(subtype, 0) + 1
        if subtype != "ASH":
            continue
        if not re.fullmatch(r"6\d{5}", code):
            raise ValueError(f"invalid ASH code from SSE quote control: {code!r}")
        codes.add(code)

    if len(codes) < 1500:
        raise RuntimeError(f"implausibly small SSE ASH control set: {len(codes)}")
    return codes, {
        "date": payload.get("date"),
        "time": payload.get("time"),
        "total_equities": total,
        "subtype_counts": subtype_counts,
        "ash_rows": len(codes),
        "classification_basis": "SSE technical specification: CPXXSubType ASH = RMB-traded main-board stock",
    }


def norm_cell(v: object) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v)).strip()


def find_header(rows: Iterable[tuple[object, ...]]) -> tuple[int, dict[str, int], list[tuple[object, ...]]]:
    materialized = list(rows)
    aliases = {
        "code": {"A股代码", "证券代码", "股票代码"},
        "board": {"板块", "所属板块", "市场板块"},
    }
    for idx, row in enumerate(materialized[:30]):
        vals = [norm_cell(v) for v in row]
        mapping: dict[str, int] = {}
        for key, names in aliases.items():
            for j, val in enumerate(vals):
                if val in names:
                    mapping[key] = j
                    break
        if "code" in mapping:
            return idx, mapping, materialized
    raise ValueError("SZSE XLSX header with A-share code not found")


def parse_szse_xlsx(raw: bytes) -> tuple[set[str], dict[str, object]]:
    # SZSE currently writes worksheet dimension=A1 despite a full XML table.
    # Normal mode parses actual cell XML; read_only mode would trust A1 and truncate.
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    diagnostics: dict[str, object] = {"sheets": wb.sheetnames}
    best_codes: set[str] = set()
    best_diag: dict[str, object] | None = None

    for ws in wb.worksheets:
        try:
            header_idx, mapping, rows = find_header(ws.iter_rows(values_only=True))
        except ValueError:
            continue
        if "board" not in mapping:
            raise ValueError(f"SZSE XLSX sheet {ws.title!r} has no explicit board column")

        codes: set[str] = set()
        board_values: set[str] = set()
        for row in rows[header_idx + 1 :]:
            if mapping["code"] >= len(row):
                continue
            code = norm_cell(row[mapping["code"]])
            if not re.fullmatch(r"\d{6}", code):
                continue
            board = norm_cell(row[mapping["board"]]) if mapping["board"] < len(row) else ""
            if board:
                board_values.add(board)
            if "主板" in board:
                codes.add(code)

        if len(codes) > len(best_codes):
            best_codes = codes
            best_diag = {
                "sheet": ws.title,
                "header_row_1based": header_idx + 1,
                "mapping": mapping,
                "board_values": sorted(board_values),
                "main_a_rows": len(codes),
            }

    if len(best_codes) < 1400:
        raise RuntimeError(f"implausibly small SZSE XLSX main-A control set: {len(best_codes)}; {diagnostics}")
    diagnostics.update(best_diag or {})
    return best_codes, diagnostics


def diff(primary: set[str], control: set[str]) -> dict[str, object]:
    only_primary = sorted(primary - control)
    only_control = sorted(control - primary)
    return {
        "primary_count": len(primary),
        "control_count": len(control),
        "set_equal": not only_primary and not only_control,
        "only_primary_count": len(only_primary),
        "only_control_count": len(only_control),
        "only_primary_sample": only_primary[:50],
        "only_control_sample": only_control[:50],
    }


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    master = root / "data/current_master"
    out_path = master / "reconciliation.json"

    sse_primary = read_primary_codes(master / "sse_main_a.csv", "SSE")
    szse_primary = read_primary_codes(master / "szse_main_a.csv", "SZSE")

    sse_raw = fetch(SSE_CONTROL, SSE_PAGE)
    szse_raw = fetch(SZSE_CONTROL, SZSE_PAGE)
    sse_control, sse_diag = parse_sse_quote(sse_raw)
    szse_control, szse_diag = parse_szse_xlsx(szse_raw)

    sse_result = diff(sse_primary, sse_control)
    szse_result = diff(szse_primary, szse_control)
    sse_result.update({
        "control_url": SSE_CONTROL,
        "control_sha256": sha256(sse_raw),
        "quote_diagnostics": sse_diag,
    })
    szse_result.update({
        "control_url": SZSE_CONTROL,
        "control_sha256": sha256(szse_raw),
        "xlsx_diagnostics": szse_diag,
    })

    reconciled = bool(sse_result["set_equal"] and szse_result["set_equal"])
    report = {
        "status": "RECONCILED" if reconciled else "MISMATCH",
        "g1_reconciled": reconciled,
        "sse": sse_result,
        "szse": szse_result,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if reconciled else 2


if __name__ == "__main__":
    sys.exit(main())
