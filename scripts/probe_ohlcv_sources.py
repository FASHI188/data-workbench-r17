#!/usr/bin/env python3
"""Probe exchange-owned historical OHLCV sources for Stage2 G3.

Diagnostics only. Tests current and delisted main-board A shares on both exchanges,
records schema/depth, and checks whether official endpoints retain delisted history.
"""
from __future__ import annotations

import hashlib
import json
import re
import time
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
SSE_REFERER = "https://www.sse.com.cn/market/price/trends/"
SZSE_REFERER = "https://www.szse.cn/market/trend/"

SSE_SAMPLES = [
    ("600000", "current_old"),
    ("603713", "current_newer"),
    ("600005", "delisted_2017"),
    ("600001", "delisted_2009"),
]
SZSE_SAMPLES = [
    ("000001", "current_old"),
    ("002031", "current_mid"),
    ("000024", "delisted_2015"),
    ("000004", "delisted_2026"),
]


def get(url: str, referer: str, attempts: int = 4) -> bytes:
    last = None
    for i in range(attempts):
        try:
            r = requests.get(
                url,
                headers={
                    "User-Agent": UA,
                    "Referer": referer,
                    "Accept": "*/*",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
                    "Connection": "close",
                },
                timeout=60,
            )
            r.raise_for_status()
            if not r.content:
                raise RuntimeError("empty response")
            return r.content
        except Exception as exc:
            last = exc
            if i + 1 < attempts:
                time.sleep(min(0.5 * (2**i), 4.0))
    raise last


def json_summary(raw: bytes):
    obj = json.loads(raw.decode("utf-8"))
    return obj


def sse_probe(code: str, label: str) -> dict:
    params = {
        "select": "date,open,high,low,close,volume,amount",
        "begin": "-5000",
        "end": "-1",
    }
    url = f"https://yunhq.sse.com.cn:32042/v1/sh1/dayk/{code}?" + urlencode(params)
    raw = get(url, SSE_REFERER)
    obj = json_summary(raw)
    rows = obj.get("kline") or []
    result = {
        "code": code,
        "label": label,
        "url": url,
        "bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "keys": list(obj.keys()),
        "begin": obj.get("begin"),
        "end": obj.get("end"),
        "row_count": len(rows),
        "first_rows": rows[:3],
        "last_rows": rows[-3:],
        "row_lengths": sorted({len(x) for x in rows if isinstance(x, list)}),
    }
    return result


def szse_market_probe(code: str, label: str) -> dict:
    params = {"cycleType": "32", "marketId": "1", "code": code}
    url = "https://www.szse.cn/api/market/ssjjhq/getHistoryData?" + urlencode(params)
    raw = get(url, SZSE_REFERER)
    obj = json_summary(raw)
    data = obj.get("data") or {}
    rows = data.get("picupdata") or [] if isinstance(data, dict) else []
    return {
        "code": code,
        "label": label,
        "url": url,
        "bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "top_keys": list(obj.keys()) if isinstance(obj, dict) else [],
        "data_keys": list(data.keys()) if isinstance(data, dict) else [],
        "row_count": len(rows) if isinstance(rows, list) else None,
        "first_rows": rows[:3] if isinstance(rows, list) else None,
        "last_rows": rows[-3:] if isinstance(rows, list) else None,
        "row_lengths": sorted({len(x) for x in rows if isinstance(x, list)}) if isinstance(rows, list) else [],
    }


def recursive_dicts(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from recursive_dicts(v)
    elif isinstance(obj, list):
        for v in obj:
            yield from recursive_dicts(v)


def szse_report_probe(code: str, label: str) -> dict:
    params = {
        "SHOWTYPE": "JSON",
        "CATALOGID": "1815_stock",
        "TABKEY": "tab1",
        "txtDMorJC": code,
        "txtBeginDate": "2015-01-01",
        "txtEndDate": "2026-07-24",
        "radioClass": "00,20,30",
        "txtSite": "all",
        "PAGENO": "1",
    }
    url = "https://www.szse.cn/api/report/ShowReport/data?" + urlencode(params)
    raw = get(url, SZSE_REFERER)
    obj = json_summary(raw)
    metadata = []
    row_samples = []
    for d in recursive_dicts(obj):
        if isinstance(d.get("metadata"), dict):
            metadata.append(d["metadata"])
        if any(k in d for k in ("zqdm", "zqjc", "jyrq", "rq", "cjje", "cjl")):
            row_samples.append(d)
    return {
        "code": code,
        "label": label,
        "url": url,
        "bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "metadata": metadata[:10],
        "row_samples": row_samples[:10],
        "top_type": type(obj).__name__,
    }


def szse_snapshot_probe(day: str) -> dict:
    params = {
        "SHOWTYPE": "xlsx",
        "CATALOGID": "1815_stock_snapshot",
        "txtBeginDate": day,
        "txtEndDate": day,
    }
    url = "https://www.szse.cn/api/report/ShowReport?" + urlencode(params)
    raw = get(url, SZSE_REFERER)
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return {
        "day": day,
        "url": url,
        "bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "first_rows": [["" if x is None else str(x) for x in row] for row in rows[:8]],
    }


def run_one(fn, *args):
    try:
        return fn(*args)
    except Exception as exc:
        return {"args": list(args), "error": f"{type(exc).__name__}: {exc}"}


def main() -> int:
    report = {
        "sse_dayk": [run_one(sse_probe, code, label) for code, label in SSE_SAMPLES],
        "szse_market_history": [run_one(szse_market_probe, code, label) for code, label in SZSE_SAMPLES],
        "szse_report_history": [run_one(szse_report_probe, code, label) for code, label in SZSE_SAMPLES],
        "szse_daily_snapshot": [
            run_one(szse_snapshot_probe, "2026-07-24"),
            run_one(szse_snapshot_probe, "2020-03-24"),
        ],
    }
    out = Path("data/ohlcv_probe")
    out.mkdir(parents=True, exist_ok=True)
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
