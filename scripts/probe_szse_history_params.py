#!/usr/bin/env python3
"""Parameter-matrix probe for SZSE CATALOGID=1815_stock historical daily data."""
from __future__ import annotations

import json
import re
import time
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
REFERER = "https://www.szse.cn/market/trend/index.html"


def get(url: str, attempts: int = 4) -> bytes:
    last = None
    for i in range(attempts):
        try:
            r = requests.get(url, headers={"User-Agent": UA, "Referer": REFERER, "Connection": "close"}, timeout=60)
            r.raise_for_status()
            if not r.content:
                raise RuntimeError("empty response")
            return r.content
        except Exception as exc:
            last = exc
            if i + 1 < attempts:
                time.sleep(min(0.5 * (2**i), 4.0))
    raise last


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def summarize_json(raw: bytes):
    obj = json.loads(raw.decode("utf-8"))
    metas = []
    rows = []
    def walk(x):
        if isinstance(x, dict):
            if isinstance(x.get("metadata"), dict): metas.append(x["metadata"])
            if "jyrq" in x and isinstance(x.get("jyrq"), str): rows.append(x)
            for v in x.values(): walk(v)
        elif isinstance(x, list):
            for v in x: walk(v)
    walk(obj)
    return {
        "metadata": [{k:m.get(k) for k in ("pageno","pagecount","recordcount","pagesize","footer")} for m in metas[:3]],
        "rows_found": len(rows),
        "first_rows": rows[:3],
        "last_rows": rows[-3:],
        "top_preview": str(obj)[:1000],
    }


def summarize_xlsx(raw: bytes):
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    ws = wb.active
    rows = [[norm(x) for x in row] for row in ws.iter_rows(values_only=True)]
    return {
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header": rows[0] if rows else [],
        "first": rows[1:4],
        "last": rows[-3:] if len(rows) > 1 else [],
    }


def query(code, begin, end, showtype, extra):
    params = {
        "SHOWTYPE": showtype,
        "CATALOGID": "1815_stock",
        "TABKEY": "tab1",
        "txtDMorJC": code,
        "txtBeginDate": begin,
        "txtEndDate": end,
        "radioClass": "00,20,30",
        "txtSite": "all",
    }
    params.update(extra)
    if showtype == "JSON":
        params.setdefault("PAGENO", "1")
        url = "https://www.szse.cn/api/report/ShowReport/data?" + urlencode(params)
    else:
        url = "https://www.szse.cn/api/report/ShowReport?" + urlencode(params)
    raw = get(url)
    return {"url": url, "bytes": len(raw), "summary": summarize_json(raw) if showtype == "JSON" else summarize_xlsx(raw)}


def run_case(case):
    try:
        return {**case, **query(case["code"], case["begin"], case["end"], case["showtype"], case.get("extra", {}))}
    except Exception as exc:
        return {**case, "error": f"{type(exc).__name__}: {exc}"}


def main():
    cases = []
    ranges = [
        ("2020-03-24","2020-03-24","single_2020"),
        ("2025-06-30","2025-06-30","single_cutoff"),
        ("2025-06-01","2025-06-30","month_cutoff"),
        ("2024-01-01","2024-12-31","year_2024"),
    ]
    extras = [
        ({}, "plain"),
        ({"txtHistoryMaxDate":"2025-06-30"}, "historymax"),
        ({"archiveDate":"2025-06-01"}, "archive_202506"),
        ({"txtHistoryMaxDate":"2025-06-30","archiveDate":"2025-06-01"}, "both"),
    ]
    for code in ("000001","000024"):
        for begin,end,rname in ranges:
            for extra,ename in extras:
                for showtype in ("JSON","xlsx"):
                    cases.append({"code":code,"begin":begin,"end":end,"range":rname,"extra_name":ename,"showtype":showtype,"extra":extra})
    report = [run_case(c) for c in cases]
    out=Path("data/ohlcv_probe")
    out.mkdir(parents=True,exist_ok=True)
    (out/"szse_history_params.json").write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf-8")
    winners=[]
    for r in report:
        s=r.get("summary",{})
        if (r["showtype"]=="xlsx" and s.get("max_row",0)>2) or (r["showtype"]=="JSON" and (s.get("rows_found",0)>0 or any((m.get("recordcount") or 0)>0 for m in s.get("metadata",[])))):
            winners.append(r)
    print(json.dumps({"winners":winners,"winner_count":len(winners),"total":len(report)},ensure_ascii=False,indent=2))


if __name__ == "__main__":
    main()
