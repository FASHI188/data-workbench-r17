#!/usr/bin/env python3
"""Probe whether SZSE CATALOGID=1815_stock can return a full-market historical day.

If successful, Stage2 G3 can acquire pre-2025-07 OHLCV by trading-day snapshots rather
than per-security queries, while retaining official row-level evidence.
"""
from __future__ import annotations

import json
import re
import time
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
REFERER = "https://www.szse.cn/market/trend/index.html"


def get(url: str, attempts: int = 5) -> bytes:
    last = None
    for i in range(attempts):
        try:
            r = requests.get(url, headers={"User-Agent": UA, "Referer": REFERER, "Connection": "close"}, timeout=90)
            r.raise_for_status()
            if not r.content:
                raise RuntimeError("empty response")
            return r.content
        except Exception as exc:
            last = exc
            if i + 1 < attempts:
                time.sleep(min(0.5 * (2**i), 4.0))
    raise last


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def xlsx_summary(raw: bytes) -> dict:
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    ws = wb.active
    rows = [[norm(x) for x in row] for row in ws.iter_rows(values_only=True)]
    data = [r for r in rows[1:] if r and re.fullmatch(r"\d{4}-\d{2}-\d{2}", r[0])]
    codes = [r[1] for r in data if len(r) > 1]
    return {
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header": rows[0] if rows else [],
        "data_rows": len(data),
        "first": data[:5],
        "last": data[-5:],
        "prefix_counts": {
            p: sum(str(c).startswith(p) for c in codes)
            for p in ("000", "001", "002", "003", "200", "300", "301")
        },
        "contains": {c: c in codes for c in ("000001", "000004", "000024", "002031", "300001", "200018")},
    }


def query(day: str, radio_class: str) -> dict:
    params = {
        "SHOWTYPE": "xlsx",
        "CATALOGID": "1815_stock",
        "TABKEY": "tab1",
        "txtDMorJC": "",
        "txtBeginDate": day,
        "txtEndDate": day,
        "radioClass": radio_class,
        "txtSite": "all",
    }
    url = "https://www.szse.cn/api/report/ShowReport?" + urlencode(params)
    raw = get(url)
    return {"day": day, "radioClass": radio_class, "url": url, "bytes": len(raw), "summary": xlsx_summary(raw)}


def run(day, rc):
    try:
        return query(day, rc)
    except Exception as exc:
        return {"day": day, "radioClass": rc, "error": f"{type(exc).__name__}: {exc}"}


def main():
    cases = []
    for day in ("2015-01-05", "2015-06-01", "2020-03-24", "2025-06-30"):
        for rc in ("00,20,30", "00", "20", "30"):
            cases.append(run(day, rc))
    out = Path("data/ohlcv_probe")
    out.mkdir(parents=True, exist_ok=True)
    (out / "szse_daily_history_bulk.json").write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(cases, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
