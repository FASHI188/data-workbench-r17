#!/usr/bin/env python3
"""Deep probe for a complete official SZSE daily OHLCV chain.

Checks:
1) CATALOGID=1815_stock XLSX through the exchange-declared historical cutoff 2025-06-30;
2) CATALOGID=1815_stock_snapshot for dates after the cutoff;
3) archiveDate behavior for older daily snapshots;
4) current and delisted securities.
"""
from __future__ import annotations

import hashlib
import json
import re
import time
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
REFERER = "https://www.szse.cn/market/trend/index.html"


def get(url: str, attempts: int = 5) -> bytes:
    last = None
    for i in range(attempts):
        try:
            r = requests.get(
                url,
                headers={"User-Agent": UA, "Referer": REFERER, "Connection": "close"},
                timeout=60,
            )
            r.raise_for_status()
            if not r.content:
                raise RuntimeError("empty response")
            return r.content
        except Exception as exc:
            last = exc
            if i + 1 < attempts:
                time.sleep(min(0.5 * (2**i), 4.0))
    raise last


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def xlsx_rows(raw: bytes):
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    ws = wb.active
    rows = [[norm(x) for x in row] for row in ws.iter_rows(values_only=True)]
    return ws.title, ws.max_row, ws.max_column, rows


def history_xlsx(code: str, begin: str, end: str) -> dict:
    params = {
        "SHOWTYPE": "xlsx",
        "CATALOGID": "1815_stock",
        "TABKEY": "tab1",
        "txtDMorJC": code,
        "txtBeginDate": begin,
        "txtEndDate": end,
        "radioClass": "00,20,30",
        "txtSite": "all",
    }
    url = "https://www.szse.cn/api/report/ShowReport?" + urlencode(params)
    raw = get(url)
    title, nr, nc, rows = xlsx_rows(raw)
    return {
        "code": code,
        "url": url,
        "bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "sheet": title,
        "max_row": nr,
        "max_column": nc,
        "header": rows[0] if rows else [],
        "first_data": rows[1:4],
        "last_data": rows[-3:] if len(rows) > 1 else [],
    }


def snapshot_xlsx(day: str, archive_date: str | None) -> dict:
    params = {
        "SHOWTYPE": "xlsx",
        "CATALOGID": "1815_stock_snapshot",
        "TABKEY": "tab1",
        "txtBeginDate": day,
        "txtEndDate": day,
    }
    if archive_date:
        params["archiveDate"] = archive_date
    url = "https://www.szse.cn/api/report/ShowReport?" + urlencode(params)
    raw = get(url)
    title, nr, nc, rows = xlsx_rows(raw)
    return {
        "day": day,
        "archive_date": archive_date,
        "url": url,
        "bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "sheet": title,
        "max_row": nr,
        "max_column": nc,
        "header": rows[0] if rows else [],
        "first_data": rows[1:3],
        "last_data": rows[-2:] if len(rows) > 1 else [],
        "contains_000001": any(row and row[1:2] == ["000001"] for row in rows[1:]) if rows else False,
    }


def run(fn, *args):
    try:
        return fn(*args)
    except Exception as exc:
        return {"args": list(args), "error": f"{type(exc).__name__}: {exc}"}


def main():
    report = {
        "history_to_2025_06_30": [
            run(history_xlsx, "000001", "2015-01-01", "2025-06-30"),
            run(history_xlsx, "002031", "2015-01-01", "2025-06-30"),
            run(history_xlsx, "000024", "2015-01-01", "2015-12-30"),
            run(history_xlsx, "000004", "2015-01-01", "2025-06-30"),
        ],
        "post_cutoff_snapshots": [
            run(snapshot_xlsx, "2025-07-01", None),
            run(snapshot_xlsx, "2025-08-01", None),
            run(snapshot_xlsx, "2025-09-22", None),
            run(snapshot_xlsx, "2026-01-05", None),
            run(snapshot_xlsx, "2026-07-24", None),
        ],
        "archive_snapshot_tests": [
            run(snapshot_xlsx, "2020-03-24", "2020-03-01"),
            run(snapshot_xlsx, "2020-03-24", "2021-06-01"),
            run(snapshot_xlsx, "2023-06-16", "2021-06-01"),
        ],
    }
    out = Path("data/ohlcv_probe")
    out.mkdir(parents=True, exist_ok=True)
    (out / "szse_depth.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
