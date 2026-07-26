#!/usr/bin/env python3
"""Stage2 G3 official daily OHLCV builder.

Production sources
------------------
SSE:
  Official SSE day-k endpoint, one request per lifecycle security. The endpoint retains
  history for delisted securities. We request up to the latest 5000 trading rows and
  keep 2015-01-01 through the G1 effective date.

SZSE:
  2015-01-01 .. 2025-06-30: official CATALOGID=1815_stock single-day XLSX with no code
  filter, which returns the full Shenzhen market for that trading day.
  2025-07-01 .. coverage end: official CATALOGID=1815_stock_snapshot single-day XLSX.

The normalized dataset is intentionally emitted as gzip CSV artifacts, not committed
as bulk market-data files to the public source repository. The repository stores code,
source digests, row counts and audit results; reproducible data artifacts are generated
by GitHub Actions.
"""
from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import re
import time
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/142 Safari/537.36"
COVERAGE_START = date(2015, 1, 1)
SZSE_HISTORY_END = date(2025, 6, 30)
SSE_REFERER = "https://www.sse.com.cn/market/price/trends/"
SZSE_REFERER = "https://www.szse.cn/market/trend/index.html"
FIELDS = ["exchange", "code", "trade_date", "open", "high", "low", "close", "volume_shares", "amount_cny"]


def sha256(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def request_bytes(url: str, referer: str, attempts: int = 6, timeout: int = 90) -> bytes:
    last: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            r = requests.get(
                url,
                headers={
                    "User-Agent": UA,
                    "Referer": referer,
                    "Accept": "*/*",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
                    "Connection": "close",
                },
                timeout=timeout,
            )
            r.raise_for_status()
            if not r.content:
                raise RuntimeError(f"empty response: {url}")
            return r.content
        except (requests.RequestException, RuntimeError) as exc:
            last = exc
            if attempt < attempts:
                time.sleep(min(0.4 * (2 ** (attempt - 1)), 6.0))
    assert last is not None
    raise last


def load_intervals(exchange: str) -> dict[str, tuple[date, date | None]]:
    path = ROOT / "data/security_lifecycle/security_intervals.csv"
    rows: dict[str, tuple[date, date | None]] = {}
    with path.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row["exchange"] != exchange:
                continue
            start = date.fromisoformat(row["listed_from"])
            end_s = (row.get("listed_to_exclusive") or "").strip()
            end = date.fromisoformat(end_s) if end_s else None
            code = row["code"]
            if code in rows:
                raise ValueError(f"multiple lifecycle intervals not supported for {exchange}:{code}")
            rows[code] = (start, end)
    if not rows:
        raise RuntimeError(f"no lifecycle intervals for {exchange}")
    return rows


def coverage_end() -> date:
    manifest = json.loads((ROOT / "data/current_master/manifest.json").read_text(encoding="utf-8"))
    raw = str(manifest.get("szse", {}).get("as_of") or "").strip()
    if raw:
        m = re.search(r"(20\d{2})[-年/.](\d{1,2})[-月/.](\d{1,2})", raw)
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return date(2026, 7, 24)


def active_on(interval: tuple[date, date | None], day: date) -> bool:
    start, end = interval
    return day >= start and (end is None or day < end)


def dec(value: object) -> Decimal:
    s = str(value).replace(",", "").strip()
    if s in {"", "-", "--", "None"}:
        raise ValueError(f"missing numeric value {value!r}")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ValueError(f"invalid numeric value {value!r}") from exc


def norm_price(v: object) -> str:
    x = dec(v)
    if x < 0:
        raise ValueError(f"negative price {v!r}")
    return format(x, "f")


def norm_nonnegative_integer(v: object, multiplier: int = 1) -> str:
    x = dec(v) * multiplier
    if x < 0 or x != x.to_integral_value():
        raise ValueError(f"expected nonnegative integer after normalization: {v!r} * {multiplier}")
    return str(int(x))


def validate_ohlc(o: Decimal, h: Decimal, l: Decimal, c: Decimal) -> None:
    if min(o, h, l, c) < 0:
        raise ValueError("negative OHLC")
    if h < max(o, l, c):
        raise ValueError(f"high invariant violated: O={o} H={h} L={l} C={c}")
    if l > min(o, h, c):
        raise ValueError(f"low invariant violated: O={o} H={h} L={l} C={c}")


def write_gzip_csv(path: Path, rows: list[dict[str, str]]) -> tuple[int, str]:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows.sort(key=lambda r: (r["trade_date"], r["code"]))
    seen: set[tuple[str, str]] = set()
    with gzip.open(path, "wt", encoding="utf-8", newline="", compresslevel=9) as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        for row in rows:
            key = (row["code"], row["trade_date"])
            if key in seen:
                raise ValueError(f"duplicate row {row['exchange']}:{key}")
            seen.add(key)
            w.writerow(row)
    return len(rows), sha256(path.read_bytes())


def sse_source_url(code: str) -> str:
    return (
        f"https://yunhq.sse.com.cn:32042/v1/sh1/dayk/{code}?"
        + urlencode({"select": "date,open,high,low,close,volume,amount", "begin": "-5000", "end": "-1"})
    )


def parse_sse_dayk(raw: bytes, code: str, interval: tuple[date, date | None], end_day: date) -> tuple[list[dict[str, str]], dict]:
    payload = json.loads(raw.decode("utf-8"))
    if str(payload.get("code")) != code:
        raise ValueError(f"SSE code mismatch: requested={code} payload={payload.get('code')}")
    source_rows = payload.get("kline") or []
    if not isinstance(source_rows, list):
        raise ValueError(f"SSE kline not list for {code}")
    rows: list[dict[str, str]] = []
    last_day: date | None = None
    for item in source_rows:
        if not isinstance(item, list) or len(item) != 7:
            raise ValueError(f"SSE malformed kline {code}: {item!r}")
        day = date.fromisoformat(f"{str(item[0])[:4]}-{str(item[0])[4:6]}-{str(item[0])[6:8]}")
        if last_day is not None and day <= last_day:
            raise ValueError(f"SSE non-increasing dates {code}: {last_day} -> {day}")
        last_day = day
        if day < COVERAGE_START or day > end_day:
            continue
        if not active_on(interval, day):
            raise ValueError(f"SSE OHLCV outside lifecycle {code} {day}")
        o, h, l, c = map(dec, item[1:5])
        validate_ohlc(o, h, l, c)
        rows.append({
            "exchange": "SSE", "code": code, "trade_date": day.isoformat(),
            "open": format(o, "f"), "high": format(h, "f"), "low": format(l, "f"), "close": format(c, "f"),
            "volume_shares": norm_nonnegative_integer(item[5]),
            "amount_cny": norm_nonnegative_integer(item[6]),
        })
    return rows, {
        "code": code, "source_total_rows": len(source_rows), "normalized_rows": len(rows),
        "source_begin": payload.get("begin"), "source_end": payload.get("end"),
        "first_normalized_date": rows[0]["trade_date"] if rows else None,
        "last_normalized_date": rows[-1]["trade_date"] if rows else None,
    }


def build_sse_shard(shard: int, shards: int, outdir: Path) -> None:
    intervals = load_intervals("SSE")
    end_day = coverage_end()
    codes = sorted(c for c, iv in intervals.items() if (iv[1] is None or iv[1] > COVERAGE_START) and iv[0] <= end_day)
    selected = [c for i, c in enumerate(codes) if i % shards == shard]
    by_year: dict[int, list[dict[str, str]]] = {}
    sources: list[dict] = []
    for idx, code in enumerate(selected, start=1):
        url = sse_source_url(code)
        raw = request_bytes(url, SSE_REFERER)
        rows, diag = parse_sse_dayk(raw, code, intervals[code], end_day)
        diag.update({"url": url, "sha256": sha256(raw), "bytes": len(raw)})
        sources.append(diag)
        for row in rows:
            by_year.setdefault(int(row["trade_date"][:4]), []).append(row)
        if idx % 50 == 0:
            print(f"SSE shard {shard}/{shards}: {idx}/{len(selected)} codes", flush=True)

    outdir.mkdir(parents=True, exist_ok=True)
    shards_meta = []
    for year, rows in sorted(by_year.items()):
        path = outdir / f"sse_{year}_shard{shard:02d}.csv.gz"
        n, digest = write_gzip_csv(path, rows)
        shards_meta.append({"year": year, "rows": n, "sha256": digest, "file": path.name})
    meta = {
        "exchange": "SSE", "shard": shard, "shards": shards, "coverage_start": COVERAGE_START.isoformat(),
        "coverage_end": end_day.isoformat(), "securities": len(selected), "sources": sources, "files": shards_meta,
    }
    (outdir / f"sse_shard{shard:02d}_sources.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def szse_url(day: date) -> str:
    catalog = "1815_stock" if day <= SZSE_HISTORY_END else "1815_stock_snapshot"
    params = {
        "SHOWTYPE": "xlsx", "CATALOGID": catalog, "TABKEY": "tab1",
        "txtBeginDate": day.isoformat(), "txtEndDate": day.isoformat(),
    }
    if catalog == "1815_stock":
        params.update({"txtDMorJC": "", "radioClass": "00,20,30", "txtSite": "all"})
    return "https://www.szse.cn/api/report/ShowReport?" + urlencode(params)


def xlsx_table(raw: bytes) -> tuple[list[str], list[list[str]]]:
    wb = load_workbook(BytesIO(raw), read_only=False, data_only=True)
    ws = wb.active
    values = [["" if v is None else re.sub(r"\s+", " ", str(v)).strip() for v in row] for row in ws.iter_rows(values_only=True)]
    if not values:
        raise ValueError("empty SZSE workbook")
    return values[0], values[1:]


def parse_szse_day(raw: bytes, day: date, intervals: dict[str, tuple[date, date | None]]) -> tuple[list[dict[str, str]], dict]:
    header, values = xlsx_table(raw)
    expected = ["交易日期", "证券代码", "证券简称", "前收", "开盘", "最高", "最低", "今收", "涨跌幅（%）", "成交量(万股)", "成交金额(万元)", "市盈率"]
    if header[:12] != expected:
        raise ValueError(f"unexpected SZSE OHLCV header {header}")
    rows: list[dict[str, str]] = []
    source_data_rows = 0
    ignored_out_of_scope = 0
    for v in values:
        if not v or v[0] == "没有找到符合条件的数据！":
            continue
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", v[0]):
            continue
        source_data_rows += 1
        source_day = date.fromisoformat(v[0])
        if source_day != day:
            raise ValueError(f"SZSE source day mismatch requested={day} row={source_day}")
        code = v[1]
        interval = intervals.get(code)
        if interval is None or not active_on(interval, day):
            ignored_out_of_scope += 1
            continue
        o, h, l, c = map(dec, (v[4], v[5], v[6], v[7]))
        validate_ohlc(o, h, l, c)
        rows.append({
            "exchange": "SZSE", "code": code, "trade_date": day.isoformat(),
            "open": format(o, "f"), "high": format(h, "f"), "low": format(l, "f"), "close": format(c, "f"),
            "volume_shares": norm_nonnegative_integer(v[9], 10_000),
            "amount_cny": norm_nonnegative_integer(v[10], 10_000),
        })
    if source_data_rows > 0 and not rows:
        raise ValueError(f"SZSE trading day {day} returned {source_data_rows} rows but zero in-scope main-A rows")
    return rows, {"source_data_rows": source_data_rows, "normalized_rows": len(rows), "ignored_out_of_scope": ignored_out_of_scope}


def build_szse_year(year: int, outdir: Path) -> None:
    intervals = load_intervals("SZSE")
    end_day = min(coverage_end(), date(year, 12, 31))
    day = max(COVERAGE_START, date(year, 1, 1))
    if day > end_day:
        raise ValueError(f"year outside coverage: {year}")
    rows: list[dict[str, str]] = []
    sources: list[dict] = []
    while day <= end_day:
        if day.weekday() < 5:
            url = szse_url(day)
            raw = request_bytes(url, SZSE_REFERER)
            parsed, diag = parse_szse_day(raw, day, intervals)
            diag.update({"day": day.isoformat(), "url": url, "sha256": sha256(raw), "bytes": len(raw)})
            sources.append(diag)
            rows.extend(parsed)
        day += timedelta(days=1)
    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / f"szse_{year}.csv.gz"
    n, digest = write_gzip_csv(path, rows)
    trading_days = sorted({r["trade_date"] for r in rows})
    meta = {
        "exchange": "SZSE", "year": year, "coverage_start": max(COVERAGE_START, date(year,1,1)).isoformat(),
        "coverage_end": end_day.isoformat(), "rows": n, "trading_days": len(trading_days),
        "first_trading_day": trading_days[0] if trading_days else None, "last_trading_day": trading_days[-1] if trading_days else None,
        "data_file": path.name, "data_sha256": digest, "sources": sources,
    }
    (outdir / f"szse_{year}_sources.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    sse = sub.add_parser("sse-shard")
    sse.add_argument("--shard", type=int, required=True)
    sse.add_argument("--shards", type=int, default=8)
    sse.add_argument("--out", default="build/g3/sse")
    sz = sub.add_parser("szse-year")
    sz.add_argument("--year", type=int, required=True)
    sz.add_argument("--out", default="build/g3/szse")
    args = ap.parse_args()
    if args.cmd == "sse-shard":
        if not (0 <= args.shard < args.shards):
            raise ValueError("invalid shard")
        build_sse_shard(args.shard, args.shards, Path(args.out))
    else:
        build_szse_year(args.year, Path(args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
